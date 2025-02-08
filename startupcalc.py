import openpyxl
from openpyxl.utils import get_column_letter

def startup_costs_calculator():
    # List of business expense categories
    categories = {
        "Equipment (Vehicle Cost, Cooking Utensils, etc) ": 0,
        "Food Supplies (How much are ingredients?)": 0,
        "Utilities": 0,
        "Licenses and Permits": 0,
        "Rent": 0,
        "Insurance": 0,
        "Lawyer and Accountant": 0,
        "Inventory": 0,
        "Employee Salaries": 0,
        "Advertising and Marketing": 0,
        "Market Research": 0,
        "Printed Marketing Materials": 0,
        "Making a Website": 0,
        "Other costs": 0
    }

    # Ask user for file name
    file_name = input("What would you like to save the file as?: ")
    
    # Ask user for number of months. make sure they entered a number
    while True:
        months = input("How many months are your startup costs based upon? ")
        if months.isdigit() and int(months) > 0:
            months = int(months)
            break
        print("Please enter a positive number.")
    
    # Ask user for costs in each category
    print("Enter estimated costs for each category:")
    for category in categories:
        while True:
            cost = input(f"How much was your {category.lower()} for {months} months? $")
            if cost.replace('.', '').isdigit():  # Check if input is a valid number before and after the decimal
                categories[category] = float(cost)
                break
            print("Please enter a valid number.")
    
    # Calculate total, monthly, and yearly costs
    total_cost = sum(categories.values())
    monthly_cost = total_cost / months
    yearly_cost = monthly_cost * 12

    # Create and fill out the xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Startup Costs Report"
    
    ws.append(["Category", "Cost ($)"])
    for category, cost in categories.items():
        ws.append([category, cost])
    ws.append(["Total Estimated Start-Up Cost", total_cost])
    ws.append(["Average Monthly Cost", monthly_cost])
    ws.append(["Projected Yearly Cost", yearly_cost])
    
    # Adjust column widths for better readability
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Save the Excel file
    wb.save(f"{file_name}.xlsx")
    print(f"\nReport saved as {file_name}.xlsx. Press Enter to exit.")
    input()

# Run the calculator when the script is executed
if __name__ == "__main__":
    startup_costs_calculator()
